#-*- coding:utf-8 -*-

import requests
import re
from html.parser import HTMLParser
from openpyxl import load_workbook
from openpyxl import Workbook
 
if __name__ == '__main__':
    wb = load_workbook(filename=r'stock.xlsx')
    sheets = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheets[0])

    # read xlsx
    rows = ws.rows
    columns = ws.columns
    index = 0

    for row in rows:
        index = index + 1
        line = [col.value for col in row]
        stock_symbol = line[0]
        
        #parse data
        if index <= 1:
            ws.cell(row=1, column=1).value = '股票代號'
            ws.cell(row=1, column=2).value = '股票名稱'
            ws.cell(row=1, column=3).value = '成交價'
            ws.cell(row=1, column=4).value = '漲跌'
            continue
        if str(stock_symbol) == 'None':
            continue
        print(stock_symbol)
        src = requests.get('https://tw.stock.yahoo.com/q/q?s='+str(stock_symbol))
        src = src.text.replace(' ', '')
        index2 = src.find('</a><br><ahref="')
        stock_name = src[index2-100:index2]
        index1 = stock_name.find('<tdalign=centerwidth=105>')
        stock_name = stock_name[index1+53:len(stock_name)]
        print(stock_name)

        index1 = src.find('加到投資組合</font><br></a></td>')
        index2 = src.find('<tdalign=centerwidth=137class="tt">')
        src = src[index1+27:index2]
        if src.find('Yahoo!奇摩股市') >= 0:
            ws.cell(row=index, column=2).value = ''
            ws.cell(row=index, column=3).value = ''
            continue
        src = src.replace('<tdalign="center"bgcolor="#FFFfff"nowrap>', '')
        src = src.replace('</td>', '')
        src = src.replace('<b>', '')
        src = src.replace('</b>', '')
        src = src.replace('<fontcolor=#009900>', '')
        src = src.replace('<fontcolor=#ff0000>', '')
        src = src.replace('<fontcolor=#000000>', '')
        src = src.splitlines()
        #
        #print('時間')
        #print(src[0])
        #    
        #print('成交')
        #print(src[1])
        #
        #print('買進')
        #print(src[2])
        #
        #print('賣出')
        #print(src[3])
        #
        #print('漲跌')
        if src[4][0] == '△':
            src[4] = src[4][1:len(src[4])]
        if src[4][0] == '▽':
            src[4] = '-'+src[4][1:len(src[4])]
        #print(src[4])
        #
        #print('張數')
        #print(src[5])
        #
        #print('昨收')
        #print(src[6])
        #
        #print('開盤')
        #print(src[7])
        #
        #print('最高')
        #print(src[8])
        #
        #print('最低')
        #print(src[9])
        
        # write xlsx
        
        ws.cell(row=index, column=2).value = stock_name
        ws.cell(row=index, column=3).value = src[1]
        ws.cell(row=index, column=4).value = src[4]

    wb.save(filename='stock.xlsx')
    
