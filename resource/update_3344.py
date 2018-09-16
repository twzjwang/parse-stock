#-*- coding:utf-8 -*-

import requests
import re
from html.parser import HTMLParser
from openpyxl import load_workbook
from openpyxl import Workbook

import re

TAG_RE = re.compile(r'<[^>]+>')

def remove_tags(text):
    return TAG_RE.sub('', text)

pages = 297

if __name__ == '__main__':
    wb = load_workbook(filename=r'funds.xlsx')
    sheets = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheets[0])
    ws.cell(row=1, column=1).value = 'code'
    ws.cell(row=1, column=2).value = 'url'
    ws.cell(row=1, column=3).value = 'title'
    ws.cell(row=1, column=4).value = 'price'
    ws.cell(row=1, column=5).value = 'K'
    ws.cell(row=1, column=6).value = 'D'

    index = 1

    for page in range(1, pages):
        src = requests.get('https://fund.cnyes.com/search/?page=' + str(page))
        src = src.text.replace(' ', '')
        index1 = src.find('<tbodyid="target_tbody"')
        index2 = src.find('</tbody></table></div></section>')
        # print(index1, index2)
        src = src[index1:index2]
        # print(src)
        while src.find('<ahref="/detail/') > -1:
            try:
                ### parse code url title
                index1 = src.find('<ahref="/detail/')
                index2 = src.find('</a></td><tdclass')
                content = src[index1 : index2]
                src = src[index2 + 5 : len(src)]
                # print(src)
                # print(index1, index2)
                # print(content)
                index1 = content.find('title')
                index2 = content.find('">')
                # print(content)
                url = 'https://fund.cnyes.com' + content[8 : index1 - 1]
                title = content[index2 + 2 : len(content)]
                code = content[50 : index1 - 1]
                index1 = code.find('/')
                index2 = code.find('/report/')
                code = code[index1 + 1 : index2]
                print(code, url , title)

                ### check 3344
                is_3344 = True
                content = requests.get(url)
                content = content.text.replace(' ', '')
                index1 = content.find("贏過N%基金")
                content = content[index1 + 6 : len(content)]
                index1 = content.find("</tr>")
                content = content[0 : index1]
                content = remove_tags(content)
                rate = content.split('%')
                print(rate)
                for i in range(0, len(rate) - 1):
                    if int(rate[1]) < 66:
                        is_3344 = False
                    if int(rate[2]) < 66:
                        is_3344 = False
                    if int(rate[4]) < 75:
                        is_3344 = False
                    if int(rate[5]) < 75:
                        is_3344 = False
                if not is_3344:
                    print("discard")
                    continue

                ### parse history to get kd
                history = requests.get('https://fund.cnyes.com/chart/chartstudy.aspx?code=' + code + '&mobile=true&country=fund&market=B&divwidth=150%25&divheight=700')
                history = history.text.replace(' ', '')
                index1 = history.find('globalData.push')
                index2 = history.find('varmyj')
                history = history[index1 + 15 : index2]
                history = history.split('globalData.push')
                for i in range(0, len(history)):
                    temp = history[i]
                    index1 = temp.find("',")
                    temp = temp[index1 + 2 : len(temp)]
                    index1 = temp.find(",")
                    temp = temp[0 : index1]
                    history[i] = float(temp)
                # print(history)
                k = 50
                d = 50
                for i in range(0, len(history)):
                    if i > 8:
                        low_9 = history[i]
                        high_9 = history[i]
                        for j in range(1, 9):
                            if history[i - j] > high_9:
                                high_9 = history[i - j]
                            if history[i - j] < low_9:
                                low_9 = history[i - j]
                        if high_9 - low_9 == 0:
                            rsv = 100
                        else:
                            rsv = 100 * (history[i] - low_9) / (high_9 - low_9)
                        k = k * 2 / 3 + rsv / 3
                        d = d * 2 / 3 + k / 3
                    else :
                        k = 50
                        d = 50
                price = history[len(history) - 1]
                print(price, k, d)

                ### press any key to continue
                #input('press any key to continue')

                ### write funds.xlsx
                index = index + 1
                ws.cell(row=index, column=1).value = code
                ws.cell(row=index, column=2).value = url
                ws.cell(row=index, column=3).value = title
                ws.cell(row=index, column=4).value = price
                ws.cell(row=index, column=5).value = k
                ws.cell(row=index, column=6).value = d
            except:
                print("except")
    wb.save(filename='funds.xlsx')
