#-*- coding:utf-8 -*-

import requests
import re
from html.parser import HTMLParser
from openpyxl import load_workbook
from openpyxl import Workbook
 
if __name__ == '__main__':
    wb = load_workbook(filename=r'funds.xlsx')
    sheets = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheets[0])

    # read xlsx
    rows = ws.rows
    columns = ws.columns
    index = 0

    for row in rows:
        index = index + 1
        line = [col.value for col in row]
        url = line[0]
        if str(url) == 'None':
            continue
        
        #parse data
        if index <= 1:
            ws.cell(row=1, column=1).value = 'url'
            ws.cell(row=1, column=2).value = 'title'
            ws.cell(row=1, column=3).value = 'date'
            ws.cell(row=1, column=4).value = 'price'
            ws.cell(row=1, column=5).value = 'diff'
            ws.cell(row=1, column=6).value = 'K'
            ws.cell(row=1, column=7).value = 'D'
            ws.cell(row=1, column=8).value = 'flash'
            continue

        index1 = url.find("detail/")
        url = url[index1 + 7 : len(url)]
        tittle, code, none = url.split('/')
        print(url, tittle, code)

        ### parse history to get kd
        history = requests.get('https://fund.cnyes.com/chart/chartstudy.aspx?code=' + code + '&mobile=true&country=fund&market=' + code[0] + '&divwidth=150%25&divheight=700')
        # print(history.text)
        history = history.text.replace(' ', '')
        index1 = history.find('globalData.push')
        index2 = history.find('varmyj')
        history = history[index1 + 15 : index2]
        history = history.split('globalData.push')
        date = history[len(history) - 1]
        index1 = date.find(",'")
        index2 = date.find("',")
        date = date[index1 + 2 : index2]
        # print(history)
        for i in range(0, len(history)):
            temp = history[i]
            index1 = temp.find("',")
            temp = temp[index1 + 2 : len(temp)]
            index1 = temp.find(",")
            temp = temp[0 : index1]
            history[i] = float(temp)
        # print(history)
        k = 50
        d = 50
        for i in range(0, len(history)):
            if i > 8:
                low_9 = history[i]
                high_9 = history[i]
                for j in range(1, 9):
                    if history[i - j] > high_9:
                        high_9 = history[i - j]
                    if history[i - j] < low_9:
                        low_9 = history[i - j]
                if high_9 - low_9 == 0:
                    rsv = 100
                else:
                    rsv = 100 * (history[i] - low_9) / (high_9 - low_9)
                k = k * 2 / 3 + rsv / 3
                d = d * 2 / 3 + k / 3
            else :
                k = 50
                d = 50
        price = history[len(history) - 1]
        diff = history[len(history) - 1] -  history[len(history) - 2]
        print(price, k, d)

        ### press any key to continue
        #input('press any key to continue')

        ### write funds.xlsx
        ws.cell(row=index, column=2).value = tittle
        ws.cell(row=index, column=3).value = date
        ws.cell(row=index, column=4).value = price
        ws.cell(row=index, column=5).value = diff
        ws.cell(row=index, column=6).value = k
        ws.cell(row=index, column=7).value = d
        ws.cell(row=index, column=8).value = "https://chart2.cnyes.com/test.html?markettype=fund!code=" + code + "!compare=false"

    wb.save(filename='funds.xlsx')
    
